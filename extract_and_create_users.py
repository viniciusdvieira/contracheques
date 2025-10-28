# extract_pages_simple_sqlite.py
import os, re, sys, sqlite3
import pdfplumber
from PyPDF2 import PdfReader, PdfWriter
from passlib.hash import bcrypt
from openpyxl import load_workbook  # <- para ler a planilha de nomes

# === CONFIG ===
PDF_PATH    = r"contracheques_allago.pdf"     # <- aponte para o seu PDF geral
OUTPUT_DIR  = r"contracheques_split"
DB_PATH     = r"contracheque.db"           # <- arquivo SQLite local
NAMES_XLSX  = r"nomematri.xlsx"           # <- A: Matricula, B: Nome, linha 1 = títulos

# Se todos são de setembro/2025, pode fixar aqui (ou None para auto)
REF_OVERRIDE = "2025-08"  # ex.: "2025-09" ou None

# Auto-detecção "SETEMBRO/2025" -> "2025-09"
MESES = {"JANEIRO":"01","FEVEREIRO":"02","MARCO":"03","MARÇO":"03","ABRIL":"04","MAIO":"05",
         "JUNHO":"06","JULHO":"07","AGOSTO":"08","SETEMBRO":"09","OUTUBRO":"10","NOVEMBRO":"11","DEZEMBRO":"12"}
REF_WORD_REGEX = re.compile(r"\b([A-ZÇÃÉÊÍÓÚÂÔÜ]+)\s*/\s*(\d{4})", re.IGNORECASE)

def ensure_dir(p):
    if p:
        os.makedirs(p, exist_ok=True)

def detect_ref(text: str):
    if REF_OVERRIDE:
        return REF_OVERRIDE
    m = REF_WORD_REGEX.search(text or "")
    if not m:
        return None
    mes = (m.group(1) or "").upper()\
        .replace("Ç","C").replace("Ã","A").replace("É","E").replace("Ê","E")\
        .replace("Í","I").replace("Ó","O").replace("Ú","U").replace("Â","A")\
        .replace("Ô","O").replace("Ü","U")
    mm = MESES.get(mes)
    return f"{m.group(2)}-{mm}" if mm else None

def detect_matricula(text: str):
    # Procura 7 dígitos nas primeiras linhas
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    for ln in lines[:5]:
        m = re.search(r"\b(\d{7})\b", ln)
        if m:
            return m.group(1)
    # fallback: procura em toda a página
    m = re.search(r"\b(\d{7})\b", text or "")
    return m.group(1) if m else None

def write_single_page(reader, page_idx, out_file):
    w = PdfWriter()
    w.add_page(reader.pages[page_idx])
    ensure_dir(os.path.dirname(out_file))
    with open(out_file, "wb") as f:
        w.write(f)

# ====== Planilha de nomes (A=matrícula, B=nome) ======
def _mat_variants(mat_raw: str):
    """Gera chaves possíveis para casar com a matrícula detectada do PDF."""
    if not mat_raw:
        return []
    digits = "".join(ch for ch in str(mat_raw) if ch.isdigit())
    if not digits:
        return []
    variants = {digits, digits.zfill(7), digits.zfill(8)}
    return list(variants)

def load_names_map(xlsx_path: str):
    """
    Lê a planilha e retorna um dicionário {matricula_normalizada: nome}.
    Aceita duplicatas; a última ocorrência prevalece.
    """
    names = {}
    if not os.path.exists(xlsx_path):
        print(f"[INFO] Planilha de nomes não encontrada: {xlsx_path} (seguindo sem nomes)")
        return names

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active  # primeira aba

    first = True
    for row in ws.iter_rows(min_row=1, values_only=True):
        if first:
            first = False
            # espera títulos: "Matricula", "Nome" (ignora o cabeçalho)
            continue
        mat, nome = (row[0], row[1]) if row else (None, None)
        if not mat or not nome:
            continue

        for key in _mat_variants(str(mat)):
            names[key] = str(nome).strip()

    wb.close()
    print(f"[INFO] Nomes carregados da planilha: {len(names)} chaves")
    return names

# ====== Camada de DB (SQLite) ======
DDL_USERS = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    matricula TEXT NOT NULL UNIQUE,
    nome TEXT,
    password_hash TEXT NOT NULL,
    must_change_password INTEGER NOT NULL DEFAULT 1
);
"""

DDL_PAYSLIPS = """
CREATE TABLE IF NOT EXISTS payslips (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    referencia TEXT NOT NULL,
    file_path TEXT NOT NULL,
    UNIQUE(user_id, referencia),
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
);
"""

def get_conn():
    newdb = not os.path.exists(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    # cria/garante schema
    conn.executescript(DDL_USERS + DDL_PAYSLIPS)
    conn.commit()
    return conn

def upsert_user(cur, matricula, nome=None):
    # Tenta encontrar usuário
    cur.execute("SELECT id, COALESCE(nome,'') FROM users WHERE matricula = ?", (matricula,))
    row = cur.fetchone()
    if row:
        uid, existing_nome = row
        # Se já existe e não tinha nome, atualiza
        if nome and not (existing_nome or "").strip():
            cur.execute("UPDATE users SET nome=? WHERE id=?", (nome.strip(), uid))
        return uid
    # Cria com senha padrão (hash bcrypt)
    password_hash = bcrypt.hash(f"agespisa{matricula}")
    cur.execute(
        "INSERT INTO users (matricula, nome, password_hash, must_change_password) VALUES (?,?,?,1)",
        (matricula, (nome or None), password_hash),
    )
    return cur.lastrowid

def upsert_payslip(cur, user_id, referencia, file_path):
    # UPSERT via UNIQUE(user_id, referencia)
    cur.execute(
        """
        INSERT INTO payslips (user_id, referencia, file_path)
        VALUES (?,?,?)
        ON CONFLICT(user_id, referencia) DO UPDATE SET file_path=excluded.file_path
        """,
        (user_id, referencia, file_path),
    )

def main():
    if not os.path.exists(PDF_PATH):
        print(f"ERRO: arquivo não encontrado: {PDF_PATH}")
        sys.exit(1)

    ensure_dir(OUTPUT_DIR)

    # Carrega mapa de nomes (se disponível)
    names_map = load_names_map(NAMES_XLSX)

    reader = PdfReader(PDF_PATH)
    with pdfplumber.open(PDF_PATH) as pdf:
        conn = get_conn()
        cur = conn.cursor()

        total = len(pdf.pages)
        ok, falhas = 0, 0

        try:
            for i in range(total):
                text = (pdf.pages[i].extract_text() or "")
                matricula = detect_matricula(text)
                referencia = detect_ref(text) or "sem-ref"

                if not matricula:
                    falhas += 1
                    print(f"[WARN] Página {i+1}: matrícula não encontrada.")
                    continue

                # Busca nome na planilha (via chaves variantes)
                nome = None
                for key in _mat_variants(matricula):
                    if key in names_map:
                        nome = names_map[key]
                        break

                # caminho: contracheques_split/<matricula>/<referencia>.pdf
                out_dir  = os.path.join(OUTPUT_DIR, matricula)
                out_file = os.path.join(out_dir, f"{referencia}.pdf")
                write_single_page(reader, i, out_file)

                uid = upsert_user(cur, matricula, nome=nome)
                upsert_payslip(cur, uid, referencia, out_file)

                ok += 1
                shown_nome = f" | nome: {nome}" if nome else ""
                print(f"OK p{i+1:03d}: {matricula}{shown_nome} | {referencia} -> {out_file}")

            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cur.close()
            conn.close()

        print(f"\nFinalizado. Sucesso: {ok} | Sem matrícula: {falhas} | Total páginas: {total}")

if __name__ == "__main__":
    main()
