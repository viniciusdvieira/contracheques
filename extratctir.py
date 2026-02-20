# extract_ir_grouped_sqlite.py
import os
import re
import sys
import sqlite3
import pdfplumber
from PyPDF2 import PdfReader, PdfWriter
from passlib.hash import bcrypt
from openpyxl import load_workbook

# === CONFIG ===
PDF_PATH    = r"InformeRendimentos.pdf"            # PDF geral do IR (várias pessoas)
OUTPUT_DIR  = r"contracheques_split"    # manter a mesma raiz
DB_PATH     = r"contracheque.db"        # mesmo SQLite existente
NAMES_XLSX  = r"servidores.xlsx"        # A=Matricula, B=Nome, C=CPF (somente dígitos)
IR_REF      = "IR-2025"                 # referência única para o documento no DB

# CPF e Nome normalmente aparecem na mesma linha: "C.P.F.: 003.464.413-07 NOME..."
CPF_REGEX_LINE = re.compile(
    r"(?:C\.?\s*P\.?\s*F\.?\s*[:\s]*)(\d{3}\.?\d{3}\.?\d{3}-?\d{2})\s+(.+)$",
    re.IGNORECASE
)
CPF_REGEX_ANY = re.compile(r"\b(\d{3}\.?\d{3}\.?\d{3}-?\d{2})\b")

# ===== Util =====
def ensure_dir(p):
    if p:
        os.makedirs(p, exist_ok=True)

def norm_digits(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())

def write_pages(reader: PdfReader, page_idxs: list[int], out_file: str):
    w = PdfWriter()
    for idx in page_idxs:
        w.add_page(reader.pages[idx])
    ensure_dir(os.path.dirname(out_file))
    with open(out_file, "wb") as f:
        w.write(f)

# ====== Planilha (A=matrícula, B=nome, C=cpf) ======
def load_servidores_map(xlsx_path: str):
    """
    Retorna:
      cpf_to_user = {cpf_digits: {"matricula": "1234567", "nome": "Fulano"}}
    """
    cpf_to_user = {}
    if not os.path.exists(xlsx_path):
        print(f"[ERRO] Planilha não encontrada: {xlsx_path}")
        return cpf_to_user

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    first = True
    for row in ws.iter_rows(min_row=1, values_only=True):
        if first:
            first = False
            continue
        matricula, nome, cpf = (row[0], row[1], row[2]) if row else (None, None, None)
        if not cpf:
            continue
        cpf_digits = norm_digits(str(cpf))
        if len(cpf_digits) != 11:
            continue

        matricula_digits = norm_digits(str(matricula)) if matricula else None
        if matricula_digits and len(matricula_digits) < 7:
            matricula_digits = matricula_digits.zfill(7)
        nome_str = str(nome).strip() if nome else None

        cpf_to_user[cpf_digits] = {"matricula": matricula_digits, "nome": nome_str}

    wb.close()
    print(f"[INFO] CPFs carregados da planilha: {len(cpf_to_user)}")
    return cpf_to_user

# ====== DB ======
DDL_USERS = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    matricula TEXT NOT NULL UNIQUE,  -- aqui é o LOGIN (matrícula antiga ou CPF novo)
    nome TEXT,
    password_hash TEXT NOT NULL,
    must_change_password INTEGER NOT NULL DEFAULT 1
);
"""

DDL_PAYSLIPS = """
CREATE TABLE IF NOT EXISTS payslips (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    referencia TEXT NOT NULL,
    file_path TEXT NOT NULL,
    UNIQUE(user_id, referencia),
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
);
"""

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    conn.executescript(DDL_USERS + DDL_PAYSLIPS)
    conn.commit()
    return conn

def find_user_by_login(cur, login: str):
    cur.execute("SELECT id, COALESCE(nome,'') FROM users WHERE matricula = ?", (login,))
    return cur.fetchone()

def create_user_with_login_and_password(cur, login: str, nome: str, plain_password: str):
    password_hash = bcrypt.hash(plain_password)
    cur.execute(
        "INSERT INTO users (matricula, nome, password_hash, must_change_password) VALUES (?,?,?,1)",
        (login, (nome or None), password_hash),
    )
    return cur.lastrowid

def upsert_user_keep_password(cur, login: str, nome: str = None):
    """
    Se existe: não mexe em senha; se não existe: cria.
    Regra para NOVO usuário (login=CPF):
      senha = "agespisa" + 3 primeiros dígitos do CPF
    """
    row = find_user_by_login(cur, login)
    if row:
        uid, existing_nome = row
        if nome and not (existing_nome or "").strip():
            cur.execute("UPDATE users SET nome=? WHERE id=?", (nome.strip(), uid))
        return uid, False

    # cria novo (CPF)
    if len(login) != 11 or not login.isdigit():
        raise ValueError(f"Login novo inválido (esperado CPF com 11 dígitos): {login}")

    senha = f"agespisa{login[:3]}"
    uid = create_user_with_login_and_password(cur, login, nome, senha)
    return uid, True

def upsert_doc(cur, user_id: int, referencia: str, file_path: str):
    cur.execute(
        """
        INSERT INTO payslips (user_id, referencia, file_path)
        VALUES (?,?,?)
        ON CONFLICT(user_id, referencia) DO UPDATE SET file_path=excluded.file_path
        """,
        (user_id, referencia, file_path),
    )

# ====== Extração do IR ======
def extract_cpf_nome_ir(text: str):
    """
    Tenta extrair CPF e Nome na linha do CPF.
    Fallback: encontra CPF em qualquer lugar; nome=None.
    """
    text = text or ""

    for ln in text.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        m = CPF_REGEX_LINE.search(ln)
        if m:
            cpf = norm_digits(m.group(1))
            nome = (m.group(2) or "").strip()
            nome = re.sub(r"\s{2,}", " ", nome)
            return cpf, nome

    m2 = CPF_REGEX_ANY.search(text)
    if m2:
        return norm_digits(m2.group(1)), None

    return None, None

def main():
    if not os.path.exists(PDF_PATH):
        print(f"ERRO: arquivo não encontrado: {PDF_PATH}")
        sys.exit(1)

    ensure_dir(OUTPUT_DIR)

    cpf_to_user = load_servidores_map(NAMES_XLSX)

    reader = PdfReader(PDF_PATH)
    with pdfplumber.open(PDF_PATH) as pdf:
        conn = get_conn()
        cur = conn.cursor()

        total = len(pdf.pages)
        ok_blocos = 0
        falhas = 0
        novos = 0

        current_cpf = None
        current_nome_ir = None
        current_pages = []

        def flush_current():
            nonlocal current_cpf, current_nome_ir, current_pages, ok_blocos, novos

            if not current_cpf or not current_pages:
                return

            # Decide login:
            # - Se CPF bate na planilha e tem matrícula -> login = matrícula (usuário já existente)
            # - Senão -> login = CPF (novo usuário)
            if current_cpf in cpf_to_user and cpf_to_user[current_cpf].get("matricula"):
                login = cpf_to_user[current_cpf]["matricula"]
                nome_final = cpf_to_user[current_cpf].get("nome") or current_nome_ir

                # Usuário deveria existir (249 já criados). Se não existir, cria com padrão antigo.
                uid_row = find_user_by_login(cur, login)
                if not uid_row:
                    uid = create_user_with_login_and_password(
                        cur,
                        login=login,
                        nome=nome_final,
                        plain_password=f"agespisa{login}"
                    )
                    novos += 1
                else:
                    uid = uid_row[0]
                    existing_nome = uid_row[1]
                    if nome_final and not (existing_nome or "").strip():
                        cur.execute("UPDATE users SET nome=? WHERE id=?", (nome_final.strip(), uid))
            else:
                login = current_cpf
                nome_final = current_nome_ir
                uid, created = upsert_user_keep_password(cur, login, nome=nome_final)
                if created:
                    novos += 1

            # Salva PDF completo do IR daquele CPF/login (todas as páginas agrupadas)
            out_dir  = os.path.join(OUTPUT_DIR, login)
            out_file = os.path.join(out_dir, f"{IR_REF}.pdf")
            write_pages(reader, current_pages, out_file)

            # Registra no DB
            upsert_doc(cur, uid, IR_REF, out_file)

            ok_blocos += 1
            shown_nome = f" | nome: {nome_final}" if nome_final else ""
            print(f"OK: cpf={current_cpf} -> login={login}{shown_nome} | paginas={len(current_pages)} | {out_file}")

            # limpa
            current_cpf = None
            current_nome_ir = None
            current_pages = []

        try:
            for i in range(total):
                text = (pdf.pages[i].extract_text() or "")
                cpf, nome_ir = extract_cpf_nome_ir(text)

                if cpf and len(cpf) == 11:
                    # Novo CPF detectado: fecha o bloco anterior (se houver)
                    if current_cpf and cpf != current_cpf:
                        flush_current()

                    # inicia/continua bloco do CPF atual
                    current_cpf = cpf
                    if nome_ir:
                        current_nome_ir = nome_ir
                    current_pages.append(i)
                else:
                    # Página sem CPF: continuação do bloco atual
                    if current_cpf:
                        current_pages.append(i)
                    else:
                        falhas += 1
                        print(f"[WARN] Página {i+1}: sem CPF e sem contexto (ignorada).")

            # flush final
            flush_current()

            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cur.close()
            conn.close()

        print(
            f"\nFinalizado. Blocos salvos: {ok_blocos} | Falhas (sem CPF/contexto): {falhas} | "
            f"Novos usuários: {novos} | Total páginas: {total}"
        )

if __name__ == "__main__":
    main()
